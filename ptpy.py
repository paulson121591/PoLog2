import os
import pickle
import re
import shutil
import sys

import pyautogui
import pyperclip
import win32com.client
from PySide2.QtCore import *
from PySide2.QtUiTools import *
from PySide2.QtWidgets import *
from openpyxl import load_workbook


class MySearch( QWidget ):
    def __init__(self):
        QWidget.__init__( self )
        layout = QFormLayout()

        self.searchBy = QComboBox()

        self.keyword = QLineEdit()

        self.list = QListWidget()

        self.label = QLabel( "Double click job number from list to copy it to your clipboard" )
        print()

        self.searchBy.addItems( ['jobnumber', 'Job Name', 'Salesman', 'Designer', 'Region', 'Street Name',
                                 'Zip Code', 'Quoted Price', 'Total Price', 'Customer Code', 'Customer Name',
                                 'Billing Street', 'Billing City', 'Billing Zip', 'Bf', 'phonenumber',
                                 'email', 'date', 'po', 'payment'] )

        self.btn = QPushButton( "Search" )

        self.btn.clicked.connect( self.searcher )
        self.list.itemDoubleClicked.connect( self.copytoboard )

        layout.addRow( self.keyword )
        layout.addRow( self.searchBy, self.btn )
        layout.addRow( self.label )
        layout.addWidget( self.list )

        self.setLayout( layout )
        self.setWindowTitle( "Search" )

    def searcher(self):
        self.list.clear()

        wd = os.getcwd()
        print( wd )
        jobs = wd + '\\Jobs\\'
        jobs = os.listdir( jobs )
        print( jobs )

        for f in jobs:
            job = pickle.load( open( wd + '\\Jobs\\' + f, 'rb' ) )
            allResults = job.get( self.searchBy.currentText(), '' )
            allResults = allResults.lower()
            key = self.keyword.text()
            key = key.lower()
            if key in str( allResults ):
                print( 'true' )
                self.list.addItem( job.get( 'jobnumber', '' ) )

    def copytoboard(self):
        jn = self.list.currentItem().text()
        pyperclip.copy( jn )


class MyPopup( QWidget ):
    def __init__(self):
        QWidget.__init__( self )
        layout = QFormLayout()
        self.btn = QPushButton( "Remove salesmen" )
        self.btn.clicked.connect( self.remove )
        wd = os.getcwd()

        salesmanList = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )

        salesmen = salesmanList.get( 'salesmen' )

        self.le = QComboBox()
        self.le.addItems( salesmen )
        layout.addRow( self.btn, self.le )
        self.btn1 = QPushButton( "get name" )

        self.setLayout( layout )
        self.setWindowTitle( "Input Dialog demo" )

    def remove(self):
        password = pyautogui.prompt( 'Password' )
        if password == '0206':
            salesmenToRemove = self.le.currentText()
            wd = os.getcwd()

            salesmanList = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )

            salesmen = salesmanList.get( 'salesmen' )

            salesmen.remove( salesmenToRemove )

            salesman = {'salesmen': salesmen}
            pickle.dump( salesman, open( wd + '\\Salesmen\\sales', "wb" ) )

            pyautogui.alert(
                'You will have to restart the program for these changes to take effect Or you can go to File -> Reload' )
            salesmanList = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )
            self.le.clear()
            self.le.addItems( salesmen )

            self.w.close()
        else:
            pyautogui.alert(
                'Sorry, Password incorrect, please contant Admin Paul Sfalanga to remove salesman (352)-460-5117' )


class Form( QObject ):
    # converting qt main file to python so python can edit and get information from it

    def __init__(self, ui_file, parent=None):
        super( Form, self ).__init__( parent )
        ui_file = QFile( ui_file )
        ui_file.open( QFile.ReadOnly )
        loader = QUiLoader()
        self.window = loader.load( ui_file )
        ui_file.close()
        wd = os.getcwd()

        salesmanList = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )
        purchasedBy = salesmanList.get( 'salesmen' )
        self.Towninfo={'Street Name':'431 Farmer Rd','City':'Townville, SC', 'Zip':'29689'}
        self.Ringinfo={'Street Name':'923 Industrial Blvd','City':'Ringgold, GA', 'Zip':'29689'}
        self.Oakinfo={'Street Name':'3703 Old Oakwood Rd','City':'Oakwood, GA', 'Zip':'30736'}
        shipTo = ['', 'Townville', 'Oakwood', 'Ringgold', 'Direct']  # TODO: update to locations
        locs = ['Town', 'Ring', 'Oak',]

        def loadPage1(self):
            self.inputLoc = self.window.findChild( QComboBox, 'inputLoc' )
            self.inputLoc.addItems(locs)
            self.inputPurchasedBy = self.window.findChild( QComboBox, 'inputPurchasedBy' )  # DONE
            self.inputPurchasedBy.addItems( purchasedBy )  # DONE
            self.inputShipto = self.window.findChild( QComboBox, 'inputShipto' )  # DONE
            self.inputShipto.addItems( shipTo )  # DONE
            self.inputRegion = self.window.findChild( QLineEdit, 'inputRegion' )  # Todo:Region to state add box in QT
            self.inputStreetName = self.window.findChild( QLineEdit, 'inputStreetName' )  # To
            self.inputZipCode = self.window.findChild( QLineEdit, 'inputZipCode' )  # To
            self.inputCustomerName = self.window.findChild( QLineEdit, 'inputCustomerName' )  # DONE
            self.inputBillingStreet = self.window.findChild( QLineEdit, 'inputBillingStreet' )  # To
            self.inputBillingCity = self.window.findChild( QLineEdit, 'inputBillingCity' )  # To
            self.inputBillingZip = self.window.findChild( QLineEdit, 'inputBillingZip' )  # To
            self.inputPhoneNumber = self.window.findChild( QLineEdit, 'inputPhoneNumber' )  # To
            self.inputEmail = self.window.findChild( QLineEdit, 'inputEmail' )  # To
            self.inputDate = self.window.findChild( QLineEdit, 'inputDate' )  # To
            self.shippingCity = self.window.findChild( QComboBox, 'shippingCity' )  # To
            # Combo Boxes
            self.inputJobNumber = self.window.findChild( QComboBox, 'inputJobNumber' )  # To
            self.comboCust = self.window.findChild( QComboBox, 'comboCust' )
            # Buttons
            self.actionRemove_Salesmen = self.window.findChild( QAction, 'actionRemove_Salesmen' )  # To
            self.actionSalesmen = self.window.findChild( QAction, 'actionSalesmen' )  # To
            self.actionDesigner = self.window.findChild( QAction, 'actionDesigner' )  # To
            self.deleteJob = self.window.findChild( QPushButton, 'deleteJob' )  # To
            buttonSave = self.window.findChild( QPushButton, 'buttonSave' )  # To
            buttonxl = self.window.findChild( QPushButton, 'buttonxl' )  # To
            buttonGenPo = self.window.findChild( QPushButton, 'buttonGenPo' )  # ToDO:GenPo
            buttonxOrder = self.window.findChild( QPushButton, 'buttonxOrder' )  # To
            saveCust = self.window.findChild( QPushButton, 'saveCust' )  # To
            self.actionReload = self.window.findChild( QAction, "actionReload" )  # To
            self.buttonSearch = self.window.findChild( QPushButton, 'buttonSearch' )  # To
            buttonSave.clicked.connect( self.save )
            buttonxl.clicked.connect( self.xl )
            buttonGenPo.clicked.connect( self.genPo)
            buttonxOrder.clicked.connect( self.order )
            self.deleteJob.clicked.connect( self.delete )
            saveCust.clicked.connect( self.saveCustAct )
            self.actionSalesmen.triggered.connect( self.addSalesmen )
            self.inputLoc.activated.connect( self.reloadprog)

            self.buttonSearch.clicked.connect( self.search )
            self.actionReload.triggered.connect( self.reloadprog )
            self.actionRemove_Salesmen.triggered.connect( self.removeSalesmen )

        def loadPage2(self):
            self.inputName = self.window.findChild( QLineEdit, 'inputName' )
            self.inputQty = self.window.findChild( QLineEdit, 'inputQty' )
            self.inputUnits = self.window.findChild( QLineEdit, 'inputUnits' )
            self.inputPricePerUnit = self.window.findChild( QLineEdit, 'inputPricePerUnit' )
            self.inputDesc = self.window.findChild( QLineEdit, 'inputDesc' )
            self.inputBL = self.window.findChild( QLineEdit, 'inputBL' )
            self.inputName_2 = self.window.findChild( QLineEdit, 'inputName_2' )
            self.inputQty_2 = self.window.findChild( QLineEdit, 'inputQty_2' )
            self.inputUnits_2 = self.window.findChild( QLineEdit, 'inputUnits_2' )
            self.inputPricePerUnit_2 = self.window.findChild( QLineEdit, 'inputPricePerUnit_2' )
            self.inputDesc_2 = self.window.findChild( QLineEdit, 'inputDesc_2' )
            self.inputBL_2 = self.window.findChild( QLineEdit, 'inputBL_2' )
            self.inputShipto.activated.connect(self.loadLocinfo)



        loadPage1(self)
        loadPage2(self)


        cwd = os.getcwd()
        pathcom = cwd + "/Cust"
        files = os.listdir( pathcom )
        for f in files:
            custs = f
            self.comboCust.addItem( custs )
        self.comboCust.setCurrentText( '' )
        self.comboCust.currentTextChanged.connect( self.custLookup )

        wd = os.getcwd()
        files = os.listdir( wd +'\\Jobs\\'+self.inputLoc.currentText() )
        for f in files:
            job = f
            self.inputJobNumber.addItem( job )
        self.inputJobNumber.setCurrentText( '' )
        self.inputJobNumber.activated.connect( self.load )

        self.window.show()

    # loads the pickle file when you type a job Number in the job number input
    def loadLocinfo(self):
        print('here')
        if self.inputShipto.currentText()=='Townville':
            self.inputStreetName.setText(self.Towninfo.get('Street Name',''))
            self.shippingCity.setCurrentText( self.Towninfo.get( 'City', '' ) )
            self.inputZipCode.setText( self.Towninfo.get( 'Zip', '' ) )
            print( 'here2' )
        if self.inputShipto.currentText()=='Oakwood':
            self.inputStreetName.setText(self.Oakinfo.get('Street Name',''))
            self.shippingCity.setCurrentText( self.Oakinfo.get( 'City', '' ) )
            self.inputZipCode.setText( self.Oakinfo.get( 'Zip', '' ) )
            print( 'here2' )
        if self.inputShipto.currentText()=='Ringgold':
            self.inputStreetName.setText(self.Ringinfo.get('Street Name',''))
            self.shippingCity.setCurrentText( self.Ringinfo.get( 'City', '' ) )
            self.inputZipCode.setText( self.Ringinfo.get( 'Zip', '' ) )
            print( 'here2' )
        if self.inputShipto.currentText()=='Direct':
            self.inputStreetName.setText('')
            self.shippingCity.setCurrentText('')
            self.inputZipCode.setText('')
            print( 'here2' )





    def load(self):
        jobNumber = self.inputJobNumber.currentText()
        wd = os.getcwd()
        files = os.listdir( wd + '\\Jobs\\'+self.inputLoc.currentText() )
        for f in files:
            job = f
            self.inputJobNumber.addItem( job )

        try:
            jobInfo = pickle.load( open( wd + '\\Jobs\\'+self.inputLoc.currentText() +'\\'+ jobNumber, "rb" ) )

            salesman = jobInfo.get( 'Salesman', '' )
            self.inputPurchasedBy.setCurrentText( salesman )
            self.shippingCity.setCurrentText(jobInfo.get('City',''))
            designer = jobInfo.get( 'Designer', '' )
            self.inputShipto.setCurrentText( designer )  # TODO: change Designer?
            streetName = jobInfo.get( 'Street Name', '' )
            self.inputStreetName.setText( streetName )
            zipCode = jobInfo.get( 'Zip Code', '' )
            self.inputZipCode.setText( zipCode )
            customerName = jobInfo.get( 'Customer Code', '' )
            self.comboCust.setCurrentText( customerName )
            customerName = jobInfo.get( 'Customer Name', '' )
            billingStreet = jobInfo.get( 'Billing Street', '' )
            billingCity = jobInfo.get( 'Billing City', '' )
            self.inputBillingStreet.setText( billingStreet )
            billingZip = jobInfo.get( 'Billing Zip', '' )
            self.inputBillingCity.setText( billingCity )
            phoneNumber = jobInfo.get( 'phonenumber', '' )
            email = jobInfo.get( 'email', '' )
            date = jobInfo.get( 'date', '' )

            #items: inputName inputQty inputUnits inputDesc inputBL inputName_2 inputQty_2 inputUnits_2 inputDesc_2 inputBL_2
            self.inputName.setText( jobInfo.get( 'inputName', '' ) )
            self.inputQty.setText( jobInfo.get( 'inputQty', '' ) )
            self.inputUnits.setText( jobInfo.get( 'inputUnits', '' ) )
            self.inputPricePerUnit.setText( jobInfo.get( 'inputPricePerUnit', '' ) )
            self.inputDesc.setText( jobInfo.get( 'inputDesc', '' ) )
            self.inputBL.setText( jobInfo.get( 'inputBL', '' ) )
            self.inputName_2.setText( jobInfo.get( 'inputName_2', '' ) )
            self.inputQty_2.setText( jobInfo.get( 'inputQty_2', '' ) )
            self.inputUnits_2.setText( jobInfo.get( 'inputUnits_2', '' ) )
            self.inputPricePerUnit_2.setText( jobInfo.get( 'inputPricePerUnit_2', '' ) )
            self.inputDesc_2.setText( jobInfo.get( 'inputDesc_2', '' ) )
            self.inputBL_2.setText( jobInfo.get( 'inputBL_2', '' ) )


            self.inputCustomerName.setText( customerName )
            self.inputBillingZip.setText( billingZip )
            self.inputPhoneNumber.setText( phoneNumber )
            self.inputEmail.setText( email )
            self.inputDate.setText( date )

            print( 'here' )

        except FileNotFoundError:
            self.inputPurchasedBy.setCurrentText( '' )
            self.inputShipto.setCurrentText( '' )
            self.inputStreetName.setText( 'No Job Found' )
            self.inputZipCode.setText('No Job Found' )
            self.inputCustomerName.setText( 'No Job Found' )
            self.inputBillingStreet.setText( 'No Job Found' )
            self.inputBillingCity.setText( 'No Job Found' )
            self.inputBillingZip.setText( 'No Job Found' )
            self.comboCust.setCurrentText( 'No Job Found' )
            self.inputPhoneNumber.setText( 'No Job Found' )
            self.inputEmail.setText( 'No Job Found' )
            self.inputDate.setText( 'No Job Found' )
            self.inputName.setText( '')
            self.inputQty.setText('' )
            self.inputUnits.setText( '')
            self.inputDesc.setText('' )
            self.inputBL.setText( '' )
            self.inputName_2.setText( '')
            self.inputQty_2.setText( '' )
            self.inputUnits_2.setText( '' )
            self.inputPricePerUnit.setText( '')
            self.inputPricePerUnit_2.setText( '' )

            self.inputDesc_2.setText( '')
            self.inputBL_2.setText( '' )

        except:
            jobInfo = pickle.load( open( wd + '\\Jobs\\'+self.inputLoc.currentText() +'\\'+ jobNumber, "rb" ) )
            salesman = jobInfo.get( 'Salesman', '' )
            self.inputPurchasedBy.setCurrentText( salesman )
            designer = jobInfo.get( 'Designer', '' )
            self.inputShipto.setCurrentText( designer )  # Todo: Change designer
            streetName = jobInfo.get( 'Street Name', '' )
            self.inputStreetName.setText( streetName )
            zipCode = jobInfo.get( 'Zip Code', '' )
            self.inputZipCode.setText( zipCode )
            customerName = jobInfo.get( 'Customer Code', '' )
            self.comboCust.setCurrentText( customerName )
            customerName = jobInfo.get( 'Customer Name', '' )
            billingStreet = jobInfo.get( 'Billing Street', '' )
            billingCity = jobInfo.get( 'Billing City', '' )
            self.inputBillingStreet.setText( billingStreet )
            billingZip = jobInfo.get( 'Billing Zip', '' )
            self.inputBillingCity.setText( billingCity )
            phoneNumber = jobInfo.get( 'phonenumber', '' )
            email = jobInfo.get( 'email', '' )
            date = jobInfo.get( 'date', '' )
            self.inputCustomerName.setText( customerName )
            self.inputBillingZip.setText( billingZip )
            self.inputPhoneNumber.setText( phoneNumber )
            self.inputEmail.setText( email )
            self.inputDate.setText( date )
            self.shippingCity.setCurrentText( jobInfo.get('City', '') )

            self.inputName.setText( jobInfo.get( 'inputName', '' ) )
            self.inputQty.setText( jobInfo.get( 'inputQty', '' ) )
            self.inputUnits.setText( jobInfo.get( 'inputUnits', '' ) )
            self.inputPricePerUnit.setText( jobInfo.get( 'inputPricePerUnit', '' ) )
            self.inputDesc.setText( jobInfo.get( 'inputDesc', '' ) )
            self.inputBL.setText( jobInfo.get( 'inputBL', '' ) )
            self.inputName_2.setText( jobInfo.get( 'inputName_2', '' ) )
            self.inputQty_2.setText( jobInfo.get( 'inputQty_2', '' ) )
            self.inputUnits_2.setText( jobInfo.get( 'inputUnits_2', '' ) )
            self.inputPricePerUnit_2.setText( jobInfo.get( 'inputPricePerUnit_2', '' ) )
            self.inputDesc_2.setText( jobInfo.get( 'inputDesc_2', '' ) )
            self.inputBL_2.setText( jobInfo.get( 'inputBL_2', '' ) )
            print( 'here2' )

    def genPo(self):
        wd = os.getcwd()
        files = os.listdir( wd +'\\Jobs\\'+self.inputLoc.currentText() )
        last = []
        for f in files:
            job = int(f)
            last.append(job)
        lastPo=last[-1]
        newPo=lastPo+1
        print(newPo)
        self.inputJobNumber.setCurrentText(str(newPo))
        self.clearAll()


    def clearAll(self):

        self.inputPurchasedBy.setCurrentText( '' )
        self.inputShipto.setCurrentText( '' )
        self.inputStreetName.setText( '' )
        self.inputZipCode.setText( '' )
        self.inputCustomerName.setText( '' )
        self.inputBillingStreet.setText( '' )
        self.inputBillingCity.setText( '' )
        self.inputBillingZip.setText( '' )
        self.comboCust.setCurrentText( '' )
        self.shippingCity.setCurrentText( '' )
        self.inputPhoneNumber.setText( '' )
        self.inputEmail.setText( '' )
        self.inputDate.setText( '' )
        self.inputName.setText( '' )
        self.inputQty.setText( '' )
        self.inputUnits.setText( '' )
        self.inputDesc.setText( '' )
        self.inputBL.setText( '' )
        self.inputName_2.setText( '' )
        self.inputQty_2.setText( '' )
        self.inputUnits_2.setText( '' )
        self.inputPricePerUnit.setText( '' )
        self.inputPricePerUnit_2.setText( '' )

        self.inputDesc_2.setText( '' )
        self.inputBL_2.setText( '' )

    # saves info to a pickle file
    def save(self):
        jobNumber = self.inputJobNumber.currentText()
        salesman = self.inputPurchasedBy.currentText()
        designer = self.inputShipto.currentText()
        street = self.inputStreetName.text()
        zipCode = self.inputZipCode.text()
        customerName = self.inputCustomerName.text()
        billingStreet = self.inputBillingStreet.text()
        billingCity = self.inputBillingCity.text()
        billingZip = self.inputBillingZip.text()
        phoneNumber = self.inputPhoneNumber.text()
        email = self.inputEmail.text()
        date = self.inputDate.text()
        inputName = self.inputName.text()
        inputQty=self.inputQty.text()
        inputUnits=self.inputUnits.text()
        inputPricePerUnit = self.inputPricePerUnit.text()

        inputDesc=self.inputDesc.text()
        inputBL=self.inputBL.text()
        inputName_2 = self.inputName_2.text()
        inputQty_2=self.inputQty_2.text()
        inputUnits_2=self.inputUnits_2.text()
        inputPricePerUnit_2 = self.inputPricePerUnit_2.text()
        inputDesc_2=self.inputDesc_2.text()
        inputBL_2=self.inputBL_2.text()
        city= self.shippingCity.currentText()




        try:

            jobInfo = {'jobnumber': jobNumber, 'Salesman': salesman, 'Designer': designer,
                        'Street Name': street,'City':city,
                       'Zip Code': zipCode,
                       'Customer Code': customerName, 'Customer Name': customerName,
                       'Billing Street': billingStreet, 'Billing City': billingCity, 'Billing Zip': billingZip,
                        'phonenumber': phoneNumber,
                       'email': email, 'date': date,'inputName':inputName,'inputQty':inputQty,'inputUnits':inputUnits,'inputPricePerUnit':inputPricePerUnit,
                       'inputDesc':inputDesc,'inputBL':inputBL,'inputName_2':inputName_2,'inputQty_2':inputQty_2,
                       'inputUnits_2':inputUnits_2,'inputPricePerUnit_2':inputPricePerUnit_2,'inputDesc_2':inputDesc_2,'inputBL_2':inputBL_2,}
            pickle.dump( jobInfo, open( jobNumber, "wb" ) )
            print(jobInfo.get('City',''))

            file = jobNumber
            shutil.move( os.getcwd() + '\\' + file, os.getcwd() + '\\Jobs\\'+self.inputLoc.currentText()+'\\' + file )
            pyautogui.alert( 'job saved' )
        except:

            pyautogui.alert( "Unknown payment type! please select paymet type." )

    # creates and xl file to print pdfs from
    def xl(self):

        try:
            jobNumber = self.inputJobNumber.currentText()
            salesman = self.inputPurchasedBy.currentText()
            designer = self.inputShipto.currentText()
            street = self.inputStreetName.text()
            city = self.shippingCity.currentText()
            zipCode = self.inputZipCode.text()

            customerName = self.inputCustomerName.text()
            billingStreet = self.inputBillingStreet.text()
            billingCity = self.inputBillingCity.text()
            billingZip = self.inputBillingZip.text()

            phoneNumber = self.inputPhoneNumber.text()
            email = self.inputEmail.text()
            date = self.inputDate.text()
            inputName = self.inputName.text()
            inputQty = self.inputQty.text()
            inputUnits = self.inputUnits.text()
            inputDesc = self.inputDesc.text()
            inputBL = self.inputBL.text()
            inputName_2 = self.inputName_2.text()
            inputQty_2 = self.inputQty_2.text()
            inputUnits_2 = self.inputUnits_2.text()
            inputDesc_2 = self.inputDesc_2.text()
            inputBL_2 = self.inputBL_2.text()
            pricePerUnit=self.inputPricePerUnit.text()
            pricePerUnit_2=self.inputPricePerUnit_2.text()


            wb = load_workbook( 'Purchase Order.xlsx' )
            ws = wb['Purchase Order Page 1']
            ws['G2'] = self.inputLoc.currentText()+jobNumber
            ws['C11'] = customerName
            ws['G7'] = date
            ws['C20'] =inputName
            ws['E12'] = street
            ws['E13'] = city+', '+zipCode

            ws['D20'] = int(inputQty)
            ws['E20'] = int(inputUnits)
            ws['F20'] = float(pricePerUnit)
            ws['C21'] = inputName_2
            ws['D21'] = int (inputQty_2)
            ws['E21'] = int (inputUnits_2)
            ws['F21'] = float (pricePerUnit_2)






            wd = os.getcwd()

            wb.save( wd + '\\xldocs\\' + jobNumber + '.xlsx' )
            pyautogui.alert( 'Data Consolidate' )
            viewXL = pyautogui.confirm( text='Would you like to view the excel now?', title='Veiw XL',
                                        buttons=['Yes', 'No'] )

            if viewXL == 'Yes':
                self.openXl()
        except:
            pyautogui.alert(
                'Data Consolidate Failed. Insure excel is closed and retry. If it still fails try to restart the program. For help please contact Paul at 352-460-5117' )

    # Creates pdf approval form and puts it in the job folder
    def order(self):

        try:

            jobNumber = self.inputJobNumber.currentText()

            o = win32com.client.Dispatch( "Excel.Application" )

            o.Visible = False

            wb_path = os.getcwd() + '\\xldocs\\' + jobNumber + '.xlsx'

            wb = o.Workbooks.Open( wb_path )

            ws_index_list = [1]  # say you want to print these sheets

            path_to_pdf = os.getcwd() + '\\xldocs\\' + jobNumber + ' Po.pdf'

            print_area = 'A1:H48'

            for index in ws_index_list:

                # off-by-one so the user can start numbering the worksheets at 1

                ws = wb.Worksheets[index - 1]

                ws.PageSetup.Zoom = False

                ws.PageSetup.FitToPagesTall = 1

                ws.PageSetup.FitToPagesWide = 1

                ws.PageSetup.PrintArea = print_area

            wb.WorkSheets( ws_index_list ).Select()

            wb.ActiveSheet.ExportAsFixedFormat( 0, path_to_pdf )
            wb.Close( True )
            pyautogui.alert( 'Order Form Created' )

            #if not os.path.exists( r'O:\Jobs\\' + jobNumber + '\Orders' ):
                #os.mkdir( r'O:\Jobs\\' + jobNumber + '\Orders' )

            #file = jobNumber + ' Order.pdf'
            #shutil.move( os.getcwd() + '\\xldocs\\' + file, 'O:\Jobs\\' + jobNumber + '\Orders\\' + file )

        except:
            pyautogui.alert( 'Unable to create pdf Have you Consolidated this job yet?' )

    # Creates email to send to Production using outlook
    def sendEmail(self):
        jobNumber = self.inputJobNumber.currentText()
        phoneNumber = self.inputPhoneNumber.text()
        date = self.inputDate.text()
        customerName = self.inputCustomerName.text()
        hangers = self.lineEditHang.text()
        beams = self.lineEditBeam.text()
        jobNotes = self.textEditJobNotes.toPlainText()
        sub = str( jobNumber ) + '-' + str(  ) + '-' + 'Order'
        const = win32com.client.constants
        olMailItem = 0x0
        obj = win32com.client.Dispatch( "Outlook.Application" )
        newMail = obj.CreateItem( olMailItem )
        newMail.Subject = sub
        newMail.BodyFormat = 2  # olFormatHTML https://msdn.microsoft.com/en-us/library/office/aa219371(v=office.11).aspx
        newMail.HTMLBody = "<HTML><BODY>This one is ready for production <br><br> Thanks,<br><br>Delivery Date:" + date + "<br>Call before Delivery " + phoneNumber + "<br><br>Hangers Ordered: " + hangers + "<br> Beams Ordered: " + beams + "<br> Special Instructions: " + jobNotes + " <br><br> Paul Sfalanga III<br>(864)772-3423</BODY></HTML>"
        newMail.To = "tstrayer@paneltruss.com; ty@paneltruss.com; mlowe@paneltruss.com; dickie@paneltruss.com; amarsingill@paneltruss.com; dlawrence@paneltruss.com; akimsey@paneltruss.com"

        newMail.display()
        # newMail.Send()

    def openJobFolder(self):
        jobNumber = self.inputJobNumber.currentText()

        path = 'O:\Jobs\\' + jobNumber
        pyperclip.copy( path )

    def custLookup(self):
        comboCust = self.window.findChild( QComboBox, 'comboCust' )
        cust = comboCust.currentText()
        try:
            custInfo = pickle.load( open( 'Cust\\' + cust, "rb" ) )
            customerName = custInfo.get( 'Customer Name', '' )
            billingStreet = custInfo.get( 'Billing Street', '' )
            billingCity = custInfo.get( 'Billing City', '' )
            billingZip = custInfo.get( 'Billing Zip', '' )
            phoneNumber = custInfo.get( 'phonenumber', '' )
            email = custInfo.get( 'email', '' )

            self.inputCustomerName.setText( customerName )
            self.inputBillingStreet.setText( billingStreet )
            self.inputBillingCity.setText( billingCity )
            self.inputBillingZip.setText( billingZip )
            self.inputPhoneNumber.setText( phoneNumber )
            self.inputEmail.setText( email )
        except:
            print( 'nocust' )
            # self.load()

    def saveCustAct(self):

        customerName = self.inputCustomerName.text()
        billingStreet = self.inputBillingStreet.text()
        billingCity = self.inputBillingCity.text()
        billingZip = self.inputBillingZip.text()
        phoneNumber = self.inputPhoneNumber.text()
        email = self.inputEmail.text()

        custInfo = {'Customer Code': customerName, 'Customer Name': customerName, 'Billing Street': billingStreet,
                    'Billing City': billingCity, 'Billing Zip': billingZip, 'phonenumber': phoneNumber, 'email': email}
        pickle.dump( custInfo, open( customerName, "wb" ) )

        file = customerName
        shutil.move( os.getcwd() + '\\' + file, os.getcwd() + '\\Cust\\' + file )

        comboCust = self.window.findChild( QComboBox, 'comboCust' )

        comboCust.addItem( file )
    def openXl(self):

        jobNumber = self.inputJobNumber.currentText()

        o = win32com.client.Dispatch( "Excel.Application" )

        o.Visible = True

        wb_path = os.getcwd() + '\\xldocs\\' + jobNumber + '.xlsx'

        wb = o.Workbooks.Open( wb_path )


    def delete(self):
        confirm = pyautogui.confirm( 'Are you sure you want to delete this job?' )
        print( confirm )
        if confirm == 'OK':
            job = self.inputJobNumber.currentText()
            file = os.getcwd() + '\\Jobs\\' + job #Todo:Redirect
            os.remove( file )

    def addSalesmen(self):
        newSalesmen = pyautogui.prompt( 'Salesman to add' )
        wd = os.getcwd()
        salesman = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )
        salesmanList = salesman.get( 'salesmen' )
        salesmanList.append( str( newSalesmen ) )
        salesman = {'salesmen': salesmanList}
        pickle.dump( salesman, open( wd + '\\Salesmen\\sales', "wb" ) )
        salesman = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )
        self.inputPurchasedBy.clear()
        salesmanList = salesman.get( 'salesmen' )
        self.inputPurchasedBy.addItems( salesmanList )
        print( salesman )

    def removeSalesmen(self):
        print( "Opening a new popup window..." )
        self.w = MyPopup()
        self.w.setGeometry( QRect( 100, 100, 400, 200 ) )
        self.w.show()

    def reloadprog(self):
        wd = os.getcwd()
        salesman = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )
        self.inputPurchasedBy.clear()
        self.inputJobNumber.clear()

        salesmanList = salesman.get( 'salesmen' )
        self.inputPurchasedBy.addItems( salesmanList )
        wd = os.getcwd()
        files = os.listdir( wd +'\\Jobs\\'+self.inputLoc.currentText() )
        for f in files:
            job = f
            self.inputJobNumber.addItem( job )
        self.inputJobNumber.setCurrentText( '' )
        self.inputJobNumber.activated.connect( self.load )

    def search(self):
        self.w = MySearch()
        self.w.setGeometry( QRect( 100, 100, 400, 200 ) )
        self.w.show()


if __name__ == '__main__':
    app = QApplication( sys.argv )
    form = Form( 'mainwindow.ui' )
    sys.exit( app.exec_() )
